#!/usr/bin/env python3
"""
Florida 2026 General Election – Candidate Scraper
Source: https://dos.elections.myflorida.com/candidates/
Runs weekly via GitHub Actions; auto-exits after STOP_DATE.
"""

import datetime
import re
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ELEC_ID     = "20261103-GEN"
LIST_URL    = "https://dos.elections.myflorida.com/candidates/CanList.asp"
DETAIL_URL  = "https://dos.elections.myflorida.com/candidates/CanDetail.asp"
STOP_DATE   = datetime.date(2026, 9, 1)
OUT_DIR     = Path("florida_reports")
MAX_WORKERS = 20

COLS = [
    "Account", "Name", "Office", "District", "Party", "Incumbent",
    "Address", "Phone", "Email", "Website",
    "Status", "Date Filed", "Date Qualified", "Method",
]

HEADERS = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) Chrome/120.0.0.0"}
DOE_PHONE = re.compile(r"850[.\-]?245[.\-]?6200")
PARTY_WORDS = {
    "Republican", "Democrat", "Democratic", "No Party Affiliation",
    "Green", "Libertarian", "Nonpartisan", "NPA", "REP", "DEM", "Independent",
}


def decode_cf_email(encoded: str) -> str:
    key = int(encoded[:2], 16)
    return "".join(chr(int(encoded[i:i+2], 16) ^ key) for i in range(2, len(encoded), 2))


def get_account_ids(session: requests.Session) -> list:
    resp = session.post(LIST_URL, data={"elecid": ELEC_ID, "GenSubmit": "View List"}, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")
    seen, accounts = set(), []
    for a in soup.find_all("a", href=re.compile(r"CanDetail\.asp\?account=\d+")):
        m = re.search(r"account=(\d+)", a["href"])
        if m:
            acct = m.group(1)
            if acct not in seen:
                seen.add(acct)
                accounts.append(acct)
    return accounts


def parse_detail(session: requests.Session, acct: str) -> dict:
    result = {col: "" for col in COLS}
    result["Account"] = acct
    try:
        resp = session.get(f"{DETAIL_URL}?account={acct}", timeout=30)
        resp.raise_for_status()
    except Exception as e:
        result["Status"] = f"ERROR: {e}"
        return result

    soup = BeautifulSoup(resp.text, "lxml")

    for span in soup.find_all("span", class_="__cf_email__"):
        cf = span.get("data-cfemail", "")
        if cf:
            span.replace_with(decode_cf_email(cf))

    nm = re.search(r"<font\s+size=\+1\s+color=.*?><b>(.*?)</b></font>", resp.text, re.DOTALL | re.I)
    if nm:
        result["Name"] = re.sub(r"\s+", " ", nm.group(1).strip())

    header_td = soup.find("td", attrs={"colspan": "4"})
    if header_td:
        parts = [p.strip() for p in header_td.get_text(separator="|", strip=True).split("|") if p.strip()]
        name_val = result["Name"]
        office_candidates = []
        for part in parts:
            if "General Election" in part or "Special Election" in part:
                continue
            if name_val and (name_val in part or part in name_val):
                continue
            if "Incumbent" in part:
                result["Incumbent"] = "Yes"
                continue
            if part in PARTY_WORDS:
                result["Party"] = part
                continue
            if re.match(r"^(District|Seat)\s+\d+", part, re.I):
                result["District"] = re.search(r"\d+", part).group()
                continue
            if re.match(r"^\d+$", part) and not result["District"]:
                result["District"] = part
                continue
            office_candidates.append(part)
        if office_candidates:
            result["Office"] = office_candidates[0]

    if not result["Party"]:
        pm = re.search(
            r"<b>\s*(Republican|Democrat(?:ic)?|No Party Affiliation|Green|Libertarian)\s*</b>",
            resp.text, re.I,
        )
        if pm:
            result["Party"] = pm.group(1)

    text_lines = soup.get_text(separator="\n", strip=True).split("\n")
    LABEL_MAP = {
        "Status:": "Status", "Date Filed:": "Date Filed",
        "Date Qualified:": "Date Qualified", "Method:": "Method",
        "Email:": "Email", "E-Mail:": "Email",
        "Website:": "Website", "Web Site:": "Website",
    }
    for i, line in enumerate(text_lines):
        if line.strip() in LABEL_MAP and i + 1 < len(text_lines):
            val = text_lines[i + 1].strip()
            field = LABEL_MAP[line.strip()]
            if field == "Status" and ":" in val:
                continue
            result[field] = val

    for td in soup.find_all("td"):
        if "Address" in td.get_text() and "Phone:" in td.get_text():
            addr_text = td.get_text(separator="|", strip=True)
            pm2 = re.search(r"Phone:\s*\|?\s*([\(\d\)\-.\s\/]{7,25})", addr_text)
            if pm2:
                ph = pm2.group(1).strip().rstrip("|").strip()
                if not DOE_PHONE.search(ph.replace(" ", "").replace(".", "").replace("-", "")):
                    result["Phone"] = ph
            am = re.search(r"Address\s*\|+(.*?)\|?\s*Phone:", addr_text, re.DOTALL)
            if am:
                addr_parts = [p.strip() for p in am.group(1).split("|") if p.strip()]
                result["Address"] = ", ".join(addr_parts).replace("\xa0", "")
            break

    for k in result:
        result[k] = result[k].replace("\xa0", " ").strip()
    return result


def main():
    today = datetime.date.today()
    if today >= STOP_DATE:
        print(f"STOP_DATE {STOP_DATE} reached - exiting.")
        return

    OUT_DIR.mkdir(exist_ok=True)
    print(f"[{today}] Starting Florida 2026 General Election candidate scrape...")

    session = requests.Session()
    session.headers.update(HEADERS)

    print("  Fetching candidate list...", end=" ", flush=True)
    accounts = get_account_ids(session)
    print(f"{len(accounts)} unique candidates found.")

    records, errors, done = [], 0, 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(parse_detail, session, acct): acct for acct in accounts}
        for future in as_completed(futures):
            rec = future.result()
            records.append(rec)
            done += 1
            if "ERROR" in rec.get("Status", ""):
                errors += 1
            if done % 100 == 0:
                print(f"  {done}/{len(accounts)} fetched ({errors} errors)...")

    records.sort(key=lambda r: (r["Office"], r["District"].zfill(4), r["Name"]))

    print(f"\nTotal records: {len(records)}, errors: {errors}")
    out_path = OUT_DIR / f"florida_candidates_{today.strftime('%Y%m%d')}.xlsx"
    write_excel(records, out_path, today)
    print(f"Excel saved: {out_path}")


def write_excel(records: list, out_path: Path, run_date: datetime.date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Florida 2026 Candidates"

    hdr_fill  = PatternFill("solid", fgColor="003087")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    alt_fill  = PatternFill("solid", fgColor="D6E4F0")
    norm_fill = PatternFill("solid", fgColor="FFFFFF")

    ws.append(COLS)
    for c in range(1, len(COLS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    for r_idx, rec in enumerate(records, start=2):
        for c_idx, col in enumerate(COLS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=rec.get(col, ""))
            cell.alignment = Alignment(vertical="top")
            cell.fill = alt_fill if r_idx % 2 == 0 else norm_fill

    widths = {
        "Account": 10, "Name": 30, "Office": 40, "District": 10,
        "Party": 14, "Incumbent": 10, "Address": 42, "Phone": 16,
        "Email": 36, "Website": 36, "Status": 14,
        "Date Filed": 14, "Date Qualified": 16, "Method": 42,
    }
    for c_idx, col in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(col, 20)

    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 44
    ws2.column_dimensions["B"].width = 12

    ws2["A1"] = "Florida 2026 General Election - Candidate Report"
    ws2["A1"].font = Font(bold=True, size=14, color="003087")
    ws2.merge_cells("A1:B1")
    ws2.append([])
    ws2.append(["Run Date",         str(run_date)])
    ws2.append(["Total Candidates", len(records)])
    ws2.append(["Records w/ Phone", sum(1 for r in records if r.get("Phone"))])
    ws2.append(["Records w/ Email", sum(1 for r in records if r.get("Email"))])
    ws2.append(["Source URL", "https://dos.elections.myflorida.com/candidates/"])
    ws2.append([])

    sec_font = Font(bold=True, size=11, color="003087")

    def section(label, key):
        ws2.append([label, "Count"])
        r = ws2.max_row
        ws2.cell(row=r, column=1).font = sec_font
        ws2.cell(row=r, column=2).font = sec_font
        for name, cnt in sorted(Counter(rec.get(key) or "N/A" for rec in records).items(), key=lambda x: -x[1]):
            ws2.append([name, cnt])
        ws2.append([])

    section("Party Breakdown",  "Party")
    section("Status Breakdown", "Status")
    section("Office Breakdown", "Office")

    wb.save(out_path)


if __name__ == "__main__":
    main()
